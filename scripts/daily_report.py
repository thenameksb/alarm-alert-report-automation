from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime, time
from datetime import (
    datetime,
    time,
    timedelta
)

# =========================================================
# TIME CONVERTER
# =========================================================

def convert_excel_datetime(value):

    if value is None:
        return None

    if isinstance(value, datetime):
        return value

    if isinstance(value, str):

        formats = [
            "%d-%m-%Y %H:%M",
            "%m/%d/%Y %I:%M:%S %p",
            "%m/%d/%Y %H:%M:%S",
            "%d/%m/%Y %H:%M:%S"
        ]

        for fmt in formats:
            try:
                return datetime.strptime(
                    value,
                    fmt
                )
            except:
                pass

    return None

# =========================================================
# CONFIG
# =========================================================

FILE_PATH = r"D:\alarm-alert-report-automation\sample_input\sample_input.xlsx" #REPLACE YOUR INPUT FILE PATH 

OUTPUT_FILE = r"D:\alarm-alert-report-automation\sample_output\sample_output.xlsx" #REPLACE YOUR OUTPUT FILE PATH 


SOURCE_SHEET = "Sheet1"
RAW_SHEET = "RAW"

START_COL_NAME = "Start Date/Time"
DURATION_COL_NAME = "Duration (HH:MM:SS)"
CUSTOM_SCALE_COL_NAME = "Custom scale"

NIGHT_START = time(21, 0)
NIGHT_END = time(5, 0)

# =========================================================
# STYLES
# =========================================================

SKY_BLUE_FONT = Font(color="00B0F0")

HEADER_FONT = Font(
    bold=True
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    start_color="92D050",
    end_color="92D050"
)
# =========================================================
# NIGHT CHECK
# =========================================================

def is_night_alert(value):

    value = convert_excel_datetime(value)

    if value is None:
        return False

    current_time = value.time()

    return (
        current_time >= NIGHT_START
        or
        current_time <= NIGHT_END
    )

# =========================================================
# LOAD WORKBOOK
# =========================================================

workbook = load_workbook(FILE_PATH)

# =========================================================
# DELETE OLD RAW
# =========================================================

if RAW_SHEET in workbook.sheetnames:
    del workbook[RAW_SHEET]

# =========================================================
# CREATE RAW
# =========================================================

raw_sheet = workbook.create_sheet(RAW_SHEET)

source_sheet = workbook[SOURCE_SHEET]

# =========================================================
# COPY DATA
# =========================================================

for row in source_sheet.iter_rows():

    for cell in row:

        raw_sheet[cell.coordinate].value = cell.value

print("✅ RAW SHEET CREATED")

# =========================================================
# FIND START COLUMN
# =========================================================

headers = [
    cell.value
    for cell in raw_sheet[1]
]

header_index = {
    header: index + 1
    for index, header in enumerate(headers)
}

START_COL = header_index[START_COL_NAME]

# =========================================================
# HEADER FORMAT
# =========================================================

for cell in raw_sheet[1]:

    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL

print("✅ HEADER FORMATTED")

# =========================================================
# NIGHT ALERT HIGHLIGHT
# =========================================================

night_count = 0

for row in range(2, raw_sheet.max_row + 1):

    start_value = raw_sheet.cell(
        row=row,
        column=START_COL
    ).value

    start_value = convert_excel_datetime(
        start_value
    )

    if start_value is None:
        continue
    


    if is_night_alert(start_value):

        for col in range(
            1,
            raw_sheet.max_column + 1
        ):

            raw_sheet.cell(
                row=row,
                column=col
            ).font = SKY_BLUE_FONT

        night_count += 1


print(
    f"🌙 NIGHT ALERTS FOUND → {night_count}"
)

print(
    "Meaning: Night time alerts highlighted in sky blue"
)

print(
    "✅ TASK 1 COMPLETED"
)

from copy import copy

# =========================================================
# TASK 2
# COPY NIGHT ALERTS FROM RAW
# TO ANALYSIS + NIGHT ALERTS RAW
# =========================================================

print("\n🚀 TASK 2 STARTED\n")

ANALYSIS_SHEET = "ANALYSIS"
NIGHT_RAW_SHEET = "NIGHT ALERTS RAW"

# =========================================================
# DELETE OLD SHEETS
# =========================================================

if ANALYSIS_SHEET in workbook.sheetnames:
    del workbook[ANALYSIS_SHEET]

if NIGHT_RAW_SHEET in workbook.sheetnames:
    del workbook[NIGHT_RAW_SHEET]

# =========================================================
# CREATE NEW SHEETS
# =========================================================

analysis_sheet = workbook.create_sheet(
    ANALYSIS_SHEET
)

night_raw_sheet = workbook.create_sheet(
    NIGHT_RAW_SHEET
)

# =========================================================
# COPY HEADER TO BOTH SHEETS
# =========================================================

for col in range(
    1,
    raw_sheet.max_column + 1
):

    source_cell = raw_sheet.cell(
        row=1,
        column=col
    )

    # ANALYSIS
    target1 = analysis_sheet.cell(
        row=1,
        column=col
    )

    target1.value = source_cell.value
    target1.font = copy(source_cell.font)
    target1.fill = copy(source_cell.fill)
    target1.border = copy(source_cell.border)
    target1.alignment = copy(source_cell.alignment)
    target1.number_format = source_cell.number_format

    # NIGHT ALERTS RAW
    target2 = night_raw_sheet.cell(
        row=1,
        column=col
    )

    target2.value = source_cell.value
    target2.font = copy(source_cell.font)
    target2.fill = copy(source_cell.fill)
    target2.border = copy(source_cell.border)
    target2.alignment = copy(source_cell.alignment)
    target2.number_format = source_cell.number_format

# =========================================================
# COPY NIGHT ALERTS
# =========================================================

analysis_row = 2
night_raw_row = 2

night_count = 0

for row in range(
    2,
    raw_sheet.max_row + 1
):

    start_value = raw_sheet.cell(
        row=row,
        column=START_COL
    ).value
    start_value = convert_excel_datetime(
        start_value
    )

    if not is_night_alert(start_value):
        continue

    night_count += 1

    for col in range(
        1,
        raw_sheet.max_column + 1
    ):

        source = raw_sheet.cell(
            row=row,
            column=col
        )

        # ANALYSIS
        dst1 = analysis_sheet.cell(
            row=analysis_row,
            column=col
        )

        dst1.value = source.value
        dst1.font = copy(source.font)
        dst1.fill = copy(source.fill)
        dst1.border = copy(source.border)
        dst1.alignment = copy(source.alignment)
        dst1.number_format = source.number_format

        # NIGHT ALERTS RAW
        dst2 = night_raw_sheet.cell(
            row=night_raw_row,
            column=col
        )

        dst2.value = source.value
        dst2.font = copy(source.font)
        dst2.fill = copy(source.fill)
        dst2.border = copy(source.border)
        dst2.alignment = copy(source.alignment)
        dst2.number_format = source.number_format

    analysis_row += 1
    night_raw_row += 1

print(
    f"🌙 NIGHT ALERTS COPIED → {night_count}"
)
print(
    "\n✅ TASK 2 COMPLETED\n"
)

from datetime import timedelta, time

def parse_duration(value):

    if value is None:
        return timedelta(0)

    if isinstance(value, timedelta):
        return value

    if isinstance(value, time):

        return timedelta(
            hours=value.hour,
            minutes=value.minute,
            seconds=value.second
        )

    if isinstance(value, str):

        try:

            h, m, s = map(
                int,
                value.split(":")
            )

            return timedelta(
                hours=h,
                minutes=m,
                seconds=s
            )

        except:
            return timedelta(0)

    return timedelta(0)

# =========================================================
# TASK 3
# MERGE NIGHT ALERTS IN ANALYSIS
# =========================================================

from copy import copy

print("\n🚀 TASK 3 STARTED\n")

MIN_NIGHT_DURATION = timedelta(minutes=5)

# =========================================================
# COLUMN MAP
# =========================================================

headers = [
    cell.value
    for cell in analysis_sheet[1]
]

header_index = {
    header: index + 1
    for index, header in enumerate(headers)
}

START_COL = header_index[
    START_COL_NAME
]

DURATION_COL = header_index[
    DURATION_COL_NAME
]

CUSTOM_SCALE_COL = header_index[
    CUSTOM_SCALE_COL_NAME
]

# =========================================================
# GROUP ALERTS
# =========================================================

night_groups = {}

for row in range(
    2,
    analysis_sheet.max_row + 1
):

    custom_scale = analysis_sheet.cell(
        row=row,
        column=CUSTOM_SCALE_COL
    ).value

    if custom_scale is None:
        continue

    start_value = analysis_sheet.cell(
        row=row,
        column=START_COL
    ).value
    start_value = convert_excel_datetime(start_value)
    
    duration_value = analysis_sheet.cell(
        row=row,
        column=DURATION_COL
    ).value

    duration_td = parse_duration(
        duration_value
    )

    row_values = []
    row_fonts = []
    row_fills = []
    row_borders = []
    row_alignments = []
    row_formats = []

    for col in range(
        1,
        analysis_sheet.max_column + 1
    ):

        cell = analysis_sheet.cell(
            row=row,
            column=col
        )

        row_values.append(cell.value)
        row_fonts.append(copy(cell.font))
        row_fills.append(copy(cell.fill))
        row_borders.append(copy(cell.border))
        row_alignments.append(
            copy(cell.alignment)
        )
        row_formats.append(
            cell.number_format
        )

    if custom_scale not in night_groups:

        night_groups[custom_scale] = []

    night_groups[
        custom_scale
    ].append({

        "start": start_value,
        "duration": duration_td,

        "values": row_values,
        "fonts": row_fonts,
        "fills": row_fills,
        "borders": row_borders,
        "alignments": row_alignments,
        "formats": row_formats

    })

print(
    f"📦 GROUPS CREATED → "
    f"{len(night_groups)}"
)

# =========================================================
# MERGE GROUPS
# =========================================================

merged_rows = []

for custom_scale, group in night_groups.items():

    master_row = min(
        group,
        key=lambda x: x["start"]
    )

    total_duration = timedelta(0)

    for item in group:

        total_duration += item[
            "duration"
        ]

    if total_duration < MIN_NIGHT_DURATION:
        continue

    master_row["duration"] = (
        total_duration
    )

    merged_rows.append(
        master_row
    )

print(
    f"🔄 MERGED ALERTS → "
    f"{len(merged_rows)}"
)

# =========================================================
# CLEAR ANALYSIS DATA
# =========================================================

if analysis_sheet.max_row > 1:

    analysis_sheet.delete_rows(
        2,
        analysis_sheet.max_row
    )

# =========================================================
# WRITE BACK
# =========================================================

write_row = 2

for item in merged_rows:

    values = item["values"]

    total_seconds = int(
        item["duration"].total_seconds()
    )

    hours = total_seconds // 3600

    minutes = (
        total_seconds % 3600
    ) // 60

    seconds = (
        total_seconds % 60
    )

    values[
        DURATION_COL - 1
    ] = (
        f"{hours:02}:{minutes:02}:{seconds:02}"
    )

    for col in range(
        1,
        len(values) + 1
    ):

        cell = analysis_sheet.cell(
            row=write_row,
            column=col
        )

        cell.value = values[
            col - 1
        ]

        cell.font = copy(
            item["fonts"][col - 1]
        )

        cell.fill = copy(
            item["fills"][col - 1]
        )

        cell.border = copy(
            item["borders"][col - 1]
        )

        cell.alignment = copy(
            item["alignments"][col - 1]
        )

        cell.number_format = (
            item["formats"][col - 1]
        )

    write_row += 1

print(
    f"✅ FINAL ANALYSIS ALERTS → "
    f"{write_row - 2}"
)

print(
    "\n✅ TASK 3 COMPLETED\n"
)

# =========================================================
# TASK 4
# DAY ALERTS
# =========================================================

from copy import copy

print("\n🚀 TASK 4 STARTED\n")

DAY_ALERTS_SHEET = "DAY ALERTS"

# =========================================================
# DELETE OLD SHEET
# =========================================================

if DAY_ALERTS_SHEET in workbook.sheetnames:
    del workbook[DAY_ALERTS_SHEET]

# =========================================================
# CREATE NEW SHEET
# =========================================================

day_sheet = workbook.create_sheet(
    DAY_ALERTS_SHEET
)

# =========================================================
# COPY HEADER
# =========================================================

for col in range(
    1,
    raw_sheet.max_column + 1
):

    source_cell = raw_sheet.cell(
        row=1,
        column=col
    )

    target = day_sheet.cell(
        row=1,
        column=col
    )

    target.value = source_cell.value
    target.font = copy(source_cell.font)
    target.fill = copy(source_cell.fill)
    target.border = copy(source_cell.border)
    target.alignment = copy(source_cell.alignment)
    target.number_format = (
        source_cell.number_format
    )

# =========================================================
# FIND COLUMN MAP
# =========================================================

headers = [
    cell.value
    for cell in day_sheet[1]
]

header_index = {
    header: index + 1
    for index, header in enumerate(headers)
}

START_COL = header_index[
    START_COL_NAME
]

DURATION_COL = header_index[
    DURATION_COL_NAME
]

CUSTOM_SCALE_COL = header_index[
    CUSTOM_SCALE_COL_NAME
]

# =========================================================
# COPY ONLY DAY ALERTS
# =========================================================

day_row = 2
day_count = 0

for row in range(
    2,
    raw_sheet.max_row + 1
):

    start_value = raw_sheet.cell(
        row=row,
        column=START_COL
    ).value


    start_value = convert_excel_datetime(
        start_value
    )


    if start_value is None:
        continue


    if is_night_alert(start_value):
        continue

    day_count += 1

    for col in range(
        1,
        raw_sheet.max_column + 1
    ):

        source = raw_sheet.cell(
            row=row,
            column=col
        )

        target = day_sheet.cell(
            row=day_row,
            column=col
        )

        target.value = source.value
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = (
            source.number_format
        )

    day_row += 1

print(
    f"☀️ DAY ALERTS COPIED → {day_count}"
)

# =========================================================
# DAY ALERTS ANALYSIS
# =========================================================

DAY_ANALYSIS_SHEET = "DAY ALERTS ANALYSIS"

# =========================================================
# DELETE OLD SHEET
# =========================================================

if DAY_ANALYSIS_SHEET in workbook.sheetnames:
    del workbook[DAY_ANALYSIS_SHEET]

# =========================================================
# CREATE SHEET
# =========================================================

day_analysis_sheet = workbook.create_sheet(
    DAY_ANALYSIS_SHEET
)

# =========================================================
# COPY HEADER
# =========================================================

for col in range(
    1,
    day_sheet.max_column + 1
):

    source_cell = day_sheet.cell(
        row=1,
        column=col
    )

    target = day_analysis_sheet.cell(
        row=1,
        column=col
    )

    target.value = source_cell.value
    target.font = copy(source_cell.font)
    target.fill = copy(source_cell.fill)
    target.border = copy(source_cell.border)
    target.alignment = copy(source_cell.alignment)
    target.number_format = (
        source_cell.number_format
    )

# =========================================================
# GROUP BY CUSTOM SCALE
# =========================================================

day_groups = {}

for row in range(
    2,
    day_sheet.max_row + 1
):

    custom_scale = day_sheet.cell(
        row=row,
        column=CUSTOM_SCALE_COL
    ).value

    if custom_scale is None:
        continue

    start_value = day_sheet.cell(
        row=row,
        column=START_COL
    ).value
    start_value = convert_excel_datetime(start_value)

    duration_value = day_sheet.cell(
        row=row,
        column=DURATION_COL
    ).value

    duration_td = parse_duration(
        duration_value
    )

    row_values = []
    row_fonts = []
    row_fills = []
    row_borders = []
    row_alignments = []
    row_formats = []

    for col in range(
        1,
        day_sheet.max_column + 1
    ):

        cell = day_sheet.cell(
            row=row,
            column=col
        )

        row_values.append(
            cell.value
        )

        row_fonts.append(
            copy(cell.font)
        )

        row_fills.append(
            copy(cell.fill)
        )

        row_borders.append(
            copy(cell.border)
        )

        row_alignments.append(
            copy(cell.alignment)
        )

        row_formats.append(
            cell.number_format
        )

    if custom_scale not in day_groups:

        day_groups[
            custom_scale
        ] = []

    day_groups[
        custom_scale
    ].append({

        "start": start_value,
        "duration": duration_td,

        "values": row_values,
        "fonts": row_fonts,
        "fills": row_fills,
        "borders": row_borders,
        "alignments": row_alignments,
        "formats": row_formats

    })

print(
    f"📦 DAY GROUPS CREATED → "
    f"{len(day_groups)}"
)

# =========================================================
# BUILD MASTER ROWS
# =========================================================

final_day_rows = []

for custom_scale, group in day_groups.items():

    total_duration = timedelta(0)

    for item in group:

        total_duration += item[
            "duration"
        ]
    # SKIP GROUPS LESS THAN 30 MINUTES

    if total_duration < timedelta(minutes=30):
        continue

    max_duration = max(
        item["duration"]
        for item in group
    )

    candidates = [

        item

        for item in group

        if item["duration"]
        == max_duration

    ]

    master_row = min(
        candidates,
        key=lambda x: x["start"]
    )

    master_row[
        "duration"
    ] = total_duration

    final_day_rows.append(
        master_row
    )

print(
    f"🔄 FINAL DAY GROUPS → "
    f"{len(final_day_rows)}"
)

# =========================================================
# WRITE RESULTS
# =========================================================

write_row = 2

for item in final_day_rows:

    values = list(
        item["values"]
    )

    total_seconds = int(
        item["duration"].total_seconds()
    )

    hours = (
        total_seconds // 3600
    )

    minutes = (
        total_seconds % 3600
    ) // 60

    seconds = (
        total_seconds % 60
    )

    values[
        DURATION_COL - 1
    ] = (
        f"{hours:02}:{minutes:02}:{seconds:02}"
    )

    for col in range(
        1,
        len(values) + 1
    ):

        cell = day_analysis_sheet.cell(
            row=write_row,
            column=col
        )

        cell.value = values[
            col - 1
        ]

        cell.font = copy(
            item["fonts"][
                col - 1
            ]
        )

        cell.fill = copy(
            item["fills"][
                col - 1
            ]
        )

        cell.border = copy(
            item["borders"][
                col - 1
            ]
        )

        cell.alignment = copy(
            item["alignments"][
                col - 1
            ]
        )

        cell.number_format = (
            item["formats"][
                col - 1
            ]
        )

    write_row += 1

print(
    f"✅ DAY ALERTS ANALYSIS ROWS → "
    f"{write_row - 2}"
)

print(
    "\n✅ TASK 4 COMPLETED\n"
)

# =========================================================
# TASK 5
# FINAL
# =========================================================

from copy import copy

print("\n🚀 TASK 5 STARTED\n")

FINAL_SHEET = "FINAL"

# =========================================================
# DELETE OLD FINAL
# =========================================================

if FINAL_SHEET in workbook.sheetnames:
    del workbook[FINAL_SHEET]

# =========================================================
# CREATE FINAL
# =========================================================

final_sheet = workbook.create_sheet(
    FINAL_SHEET
)

# =========================================================
# COPY HEADER FROM ANALYSIS
# =========================================================

for col in range(
    1,
    analysis_sheet.max_column + 1
):

    source_cell = analysis_sheet.cell(
        row=1,
        column=col
    )

    target = final_sheet.cell(
        row=1,
        column=col
    )

    target.value = source_cell.value
    target.font = copy(source_cell.font)
    target.fill = copy(source_cell.fill)
    target.border = copy(source_cell.border)
    target.alignment = copy(source_cell.alignment)
    target.number_format = (
        source_cell.number_format
    )

# =========================================================
# COLUMN MAP
# =========================================================

headers = [
    cell.value
    for cell in final_sheet[1]
]

header_index = {
    header: index + 1
    for index, header in enumerate(headers)
}

CUSTOM_SCALE_COL = header_index[
    CUSTOM_SCALE_COL_NAME
]

DURATION_COL = header_index[
    DURATION_COL_NAME
]

# =========================================================
# HELPER
# =========================================================

def duration_to_timedelta(value):

    return parse_duration(value)

# =========================================================
# LOAD NIGHT ALERTS
# =========================================================

night_data = {}

for row in range(
    2,
    analysis_sheet.max_row + 1
):

    custom_scale = analysis_sheet.cell(
        row=row,
        column=CUSTOM_SCALE_COL
    ).value

    if custom_scale is None:
        continue

    values = []
    fonts = []
    fills = []
    borders = []
    alignments = []
    formats = []

    for col in range(
        1,
        analysis_sheet.max_column + 1
    ):

        cell = analysis_sheet.cell(
            row=row,
            column=col
        )

        values.append(cell.value)
        fonts.append(copy(cell.font))
        fills.append(copy(cell.fill))
        borders.append(copy(cell.border))
        alignments.append(
            copy(cell.alignment)
        )
        formats.append(
            cell.number_format
        )

    duration_td = duration_to_timedelta(
        analysis_sheet.cell(
            row=row,
            column=DURATION_COL
        ).value
    )

    night_data[
        custom_scale
    ] = {

        "source": "NIGHT",
        "duration": duration_td,

        "values": values,
        "fonts": fonts,
        "fills": fills,
        "borders": borders,
        "alignments": alignments,
        "formats": formats
    }

# =========================================================
# LOAD DAY ALERTS ANALYSIS
# =========================================================

day_data = {}

for row in range(
    2,
    day_analysis_sheet.max_row + 1
):

    custom_scale = day_analysis_sheet.cell(
        row=row,
        column=CUSTOM_SCALE_COL
    ).value

    if custom_scale is None:
        continue

    values = []
    fonts = []
    fills = []
    borders = []
    alignments = []
    formats = []

    for col in range(
        1,
        day_analysis_sheet.max_column + 1
    ):

        cell = day_analysis_sheet.cell(
            row=row,
            column=col
        )

        values.append(cell.value)
        fonts.append(copy(cell.font))
        fills.append(copy(cell.fill))
        borders.append(copy(cell.border))
        alignments.append(
            copy(cell.alignment)
        )
        formats.append(
            cell.number_format
        )

    duration_td = duration_to_timedelta(
        day_analysis_sheet.cell(
            row=row,
            column=DURATION_COL
        ).value
    )

    day_data[
        custom_scale
    ] = {

        "source": "DAY",
        "duration": duration_td,

        "values": values,
        "fonts": fonts,
        "fills": fills,
        "borders": borders,
        "alignments": alignments,
        "formats": formats
    }

# =========================================================
# MERGE RULES
# =========================================================

all_custom_scales = set(
    list(night_data.keys())
    +
    list(day_data.keys())
)

final_rows = []

for custom_scale in all_custom_scales:

    night_row = night_data.get(
        custom_scale
    )

    day_row = day_data.get(
        custom_scale
    )

    # ONLY NIGHT

    if night_row and not day_row:

        final_rows.append(
            night_row
        )

        continue

    # ONLY DAY

    if day_row and not night_row:

        final_rows.append(
            day_row
        )

        continue

    # BOTH EXIST

    if night_row and day_row:

        if night_row[
            "duration"
        ] >= timedelta(
            minutes=5
        ):

            final_rows.append(
                night_row
            )

        else:

            final_rows.append(
                day_row
            )

# =========================================================
# WRITE FINAL
# =========================================================

write_row = 2

for item in final_rows:

    for col in range(
        1,
        len(item["values"]) + 1
    ):

        cell = final_sheet.cell(
            row=write_row,
            column=col
        )

        cell.value = item[
            "values"
        ][col - 1]

        cell.font = copy(
            item["fonts"][col - 1]
        )

        cell.fill = copy(
            item["fills"][col - 1]
        )

        cell.border = copy(
            item["borders"][col - 1]
        )

        cell.alignment = copy(
            item["alignments"][col - 1]
        )

        cell.number_format = (
            item["formats"][col - 1]
        )

    write_row += 1

print(
    f"✅ FINAL ROWS → "
    f"{write_row - 2}"
)

print(
    "\n✅ TASK 5 COMPLETED\n"
)

# =========================================================
# TASK 6
# FORCED DAY ALERT FONT
# =========================================================

print("\n🚀 TASK 6 STARTED\n")

FORCED_DAY_SCALES = {

    
    "CH101.987",
    "CH306.993",
    "CH158.453",
    "CH300.454",
    "CH111.456",
    "CH009.429",
    "CH089.615",
    "CH260.634",   

}

# =========================================================
# BUILD DAY ALERT LOOKUP
# =========================================================

day_font_lookup = {}

for row in range(
    2,
    day_sheet.max_row + 1
):

    custom_scale = day_sheet.cell(
        row=row,
        column=CUSTOM_SCALE_COL
    ).value

    if custom_scale is None:
        continue

    fonts = []

    for col in range(
        1,
        day_sheet.max_column + 1
    ):

        fonts.append(
            copy(
                day_analysis_sheet.cell(
                    row=row,
                    column=col
                ).font
            )
        )

    day_font_lookup[
        custom_scale
    ] = fonts

# =========================================================
# APPLY DAY FONT TO FINAL
# =========================================================

updated_count = 0

for row in range(
    2,
    final_sheet.max_row + 1
):

    custom_scale = final_sheet.cell(
        row=row,
        column=CUSTOM_SCALE_COL
    ).value

    if custom_scale not in FORCED_DAY_SCALES:
        continue

    if custom_scale not in day_font_lookup:
        print(
            f"{custom_scale} SKIPPED XX"
        )

        continue


    print(
        f"{custom_scale} FORCED TO DAY ALERT ✅"
    )
        

    for col in range(
        1,
        final_sheet.max_column + 1
    ):

        final_sheet.cell(
            row=row,
            column=col
        ).font = copy(
            day_font_lookup[
                custom_scale
            ][col - 1]
        )

    updated_count += 1

    

print(
    f"🎨 DAY FONT APPLIED → "
    f"{updated_count}"
)

print(
    "\n✅ TASK 6 COMPLETED\n"
)

# =========================================================
# DAILY OPERATOR ALERTS
# =========================================================

OPERATOR_ALERT_SCALES = {

    "CH101.987",
    "CH306.993",
    "CH158.453",
    "CH300.454",
    "CH111.456",
    "CH009.429",
    "CH089.615",
    "CH260.634",  
   
    
}

# =========================================================
# TASK 7
# OPERATOR ALERTS (DAY ALERTS ONLY)
# =========================================================

print("\n🚀 TASK 7 STARTED\n")

OPERATOR_ALERTS_SHEET = "OPERATOR ALERTS"

# =========================================================
# DELETE OLD SHEET
# =========================================================

if OPERATOR_ALERTS_SHEET in workbook.sheetnames:
    del workbook[OPERATOR_ALERTS_SHEET]

# =========================================================
# CREATE SHEET
# =========================================================

operator_sheet = workbook.create_sheet(
    OPERATOR_ALERTS_SHEET
)

# =========================================================
# COPY HEADER
# =========================================================

for col in range(
    1,
    day_sheet.max_column + 1
):

    source_cell = day_sheet.cell(
        row=1,
        column=col
    )

    target = operator_sheet.cell(
        row=1,
        column=col
    )

    target.value = source_cell.value
    target.font = copy(source_cell.font)
    target.fill = copy(source_cell.fill)
    target.border = copy(source_cell.border)
    target.alignment = copy(source_cell.alignment)
    target.number_format = (
        source_cell.number_format
    )

# =========================================================
# FIND BEST ROW IN DAY ALERTS
# =========================================================

def find_best_operator_row(
    custom_scale
):

    matches = []

    for row in range(
        2,
        day_sheet.max_row + 1
    ):

        scale_value = day_sheet.cell(
            row=row,
            column=CUSTOM_SCALE_COL
        ).value

        if scale_value is None:
            continue

        scale_value = str(
            scale_value
        ).strip()
        
        if scale_value.upper() != custom_scale.upper():
            continue

        

        start_value = day_sheet.cell(
            row=row,
            column=START_COL
        ).value

        duration_value = day_sheet.cell(
            row=row,
            column=DURATION_COL
        ).value

        duration_td = parse_duration(
            duration_value
        )

        matches.append({

            "row": row,
            "start": start_value,
            "duration": duration_td

        })

    if not matches:
        return None
    
    if matches:

        print(
            f"\nFOUND      :- {custom_scale} MATCHES: {len(matches)}"
        )

        for item in matches:

            print(
                f"row        :- {item['row']}"
            )

            start_value = item["start"]

            if hasattr(start_value, "strftime"):

                start_text = start_value.strftime(
                    "%d/%m/%Y %H:%M"
                )

            else:

                start_text = str(
                    start_value
                )


            print(
                f"start time :- {start_text}"
            )

            print(
                f"duration   :- {str(item['duration'])}"
            )

            print()

    else:

        print(
            f"\nNOT FOUND: {custom_scale} XX"
        )

        return None

    max_duration = max(
        item["duration"]
        for item in matches
    )

    candidates = [

        item

        for item in matches

        if item["duration"]
        == max_duration

    ]

    best_match = min(
        candidates,
        key=lambda x: x["start"]
    )

    return best_match["row"]

# =========================================================
# BUILD OPERATOR ALERTS
# =========================================================

write_row = 2
operator_count = 0

for custom_scale in OPERATOR_ALERT_SCALES:

    custom_scale = str(
        custom_scale
    ).strip()

    best_row = find_best_operator_row(
        custom_scale
    )

    if best_row is None:
        print(
            "x NOT FOUND:",
            custom_scale
        )
        continue
        

    for col in range(
        1,
        day_sheet.max_column + 1
    ):

        source = day_sheet.cell(
            row=best_row,
            column=col
        )

        target = operator_sheet.cell(
            row=write_row,
            column=col
        )

        target.value = source.value
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = (
            source.number_format
        )

    write_row += 1
    operator_count += 1

print(
    f"👷 OPERATOR ALERTS SHEET CREATED : {operator_count}"
)

print(
    "Meaning: Selected best alert for each operator scale"
)
print(
    "\n✅ TASK 7 COMPLETED\n"
)

# =========================================================
# TASK 8
# MERGE OPERATOR ALERTS INTO FINAL
# =========================================================

print("\n🚀 TASK 8 STARTED\n")

GREEN_FONT = Font(
    color="00B050"
)

# =========================================================
# BUILD ANALYSIS LOOKUP
# =========================================================

analysis_custom_scales = set()

for row in range(
    2,
    analysis_sheet.max_row + 1
):

    custom_scale = analysis_sheet.cell(
        row=row,
        column=CUSTOM_SCALE_COL
    ).value

    if custom_scale is None:
        continue

    analysis_custom_scales.add(
        str(custom_scale).strip()
    )

# =========================================================
# BUILD FINAL LOOKUP
# =========================================================

final_lookup = {}

for row in range(
    2,
    final_sheet.max_row + 1
):

    custom_scale = final_sheet.cell(
        row=row,
        column=CUSTOM_SCALE_COL
    ).value

    if custom_scale is None:
        continue

    final_lookup[
        str(custom_scale).strip()
    ] = row

# =========================================================
# PROCESS OPERATOR ALERTS
# =========================================================

updated_rows = 0
added_rows = 0

for operator_row in range(
    2,
    operator_sheet.max_row + 1
):

    custom_scale = operator_sheet.cell(
        row=operator_row,
        column=CUSTOM_SCALE_COL
    ).value

    if custom_scale is None:
        continue

    custom_scale = str(
        custom_scale
    ).strip()
    print(
        f"\nPROCESSING OPERATOR:{custom_scale}",
        
    )

    # =====================================================
    # EXISTS IN FINAL
    # =====================================================

    if custom_scale in final_lookup:

        final_row = final_lookup[
            custom_scale
        ]

        # NIGHT ALERT
        if custom_scale in analysis_custom_scales:
            continue

        # DAY ALERT
        for col in range(
            1,
            final_sheet.max_column + 1
        ):

            final_sheet.cell(
                row=final_row,
                column=col
            ).font = copy(
                SKY_BLUE_FONT
            )

        final_sheet.cell(
            row=final_row,
            column=DURATION_COL
        ).font = copy(
            GREEN_FONT
        )
        print(
            f"{custom_scale} EXIST IN FINAL SHEET ✅✅"
        )
        
        updated_rows += 1

    # =====================================================
    # NOT FOUND IN FINAL
    # =====================================================

    else:

        target_row = (
            final_sheet.max_row + 1
        )

        for col in range(
            1,
            operator_sheet.max_column + 1
        ):

            source = operator_sheet.cell(
                row=operator_row,
                column=col
            )

            target = final_sheet.cell(
                row=target_row,
                column=col
            )

            target.value = source.value

            target.font = copy(
                source.font
            )

            target.fill = copy(
                source.fill
            )

            target.border = copy(
                source.border
            )

            target.alignment = copy(
                source.alignment
            )

            target.number_format = (
                source.number_format
            )

        # APPLY SKY BLUE FONT

        for col in range(
            1,
            final_sheet.max_column + 1
        ):

            final_sheet.cell(
                row=target_row,
                column=col
            ).font = copy(
                SKY_BLUE_FONT
            )

        # APPLY GREEN DURATION FONT

        final_sheet.cell(
            row=target_row,
            column=DURATION_COL
        ).font = copy(
            GREEN_FONT
        )

        added_rows += 1

print(
    f"\n🎨 UPDATED DAY ALERT ROWS → {updated_rows}"
)

print(
    f"{updated_rows} existing CUSTOM SCALE values "
    f"found in FINAL sheet "
    f"(duration green updated)"
)


print(
    f"\n➕ OPERATOR ROWS ADDED → {added_rows}"
)

print(
    f"{added_rows} NEW CUSTOM SCALE VALUES "
    f"ADDED TO FINAL SHEET"
)

print(
    "\n✅ TASK 8 COMPLETED\n"
)
# =========================================================
# TASK 9
# MAP LOCATION COLUMN
# =========================================================

print("\n🚀 TASK 9 STARTED\n")

# =========================================================
# FIND HEADER COLUMNS DYNAMICALLY
# =========================================================

headers = [
    cell.value
    for cell in final_sheet[1]
]

header_index = {
    header: index + 1
    for index, header in enumerate(headers)
}

CUSTOM_SCALE_COL = header_index[
    CUSTOM_SCALE_COL_NAME
]

LATITUDE_COL = header_index[
    "Latitude"
]

LONGITUDE_COL = header_index[
    "Longitude"
]

# =========================================================
# REMOVE OLD MAP LOCATION COLUMN IF EXISTS
# =========================================================

if "Map Location" in headers:

    old_map_col = header_index[
        "Map Location"
    ]

    final_sheet.delete_cols(
        old_map_col
    )

    headers = [
        cell.value
        for cell in final_sheet[1]
    ]

    header_index = {
        header: index + 1
        for index, header in enumerate(headers)
    }

    CUSTOM_SCALE_COL = header_index[
        CUSTOM_SCALE_COL_NAME
    ]

    LATITUDE_COL = header_index[
        "Latitude"
    ]

    LONGITUDE_COL = header_index[
        "Longitude"
    ]

# =========================================================
# INSERT NEW MAP LOCATION COLUMN
# =========================================================

MAP_LOCATION_COL = (
    CUSTOM_SCALE_COL + 1
)

final_sheet.insert_cols(
    MAP_LOCATION_COL
)

# =========================================================
# HEADER
# =========================================================

header_cell = final_sheet.cell(
    row=1,
    column=MAP_LOCATION_COL
)

header_cell.value = (
    "Map Location"
)

header_cell.font = copy(
    final_sheet.cell(
        row=1,
        column=CUSTOM_SCALE_COL
    ).font
)

header_cell.fill = copy(
    final_sheet.cell(
        row=1,
        column=CUSTOM_SCALE_COL
    ).fill
)

header_cell.border = copy(
    final_sheet.cell(
        row=1,
        column=CUSTOM_SCALE_COL
    ).border
)

header_cell.alignment = copy(
    final_sheet.cell(
        row=1,
        column=CUSTOM_SCALE_COL
    ).alignment
)

# =========================================================
# REFRESH COLUMN POSITIONS
# =========================================================

headers = [
    cell.value
    for cell in final_sheet[1]
]

header_index = {
    header: index + 1
    for index, header in enumerate(headers)
}

CUSTOM_SCALE_COL = header_index[
    CUSTOM_SCALE_COL_NAME
]

MAP_LOCATION_COL = header_index[
    "Map Location"
]

LATITUDE_COL = header_index[
    "Latitude"
]

LONGITUDE_COL = header_index[
    "Longitude"
]

# =========================================================
# POPULATE MAP LINKS
# =========================================================

for row in range(
    2,
    final_sheet.max_row + 1
):

    latitude = final_sheet.cell(
        row=row,
        column=LATITUDE_COL
    ).value

    longitude = final_sheet.cell(
        row=row,
        column=LONGITUDE_COL
    ).value

    map_url = (
        f"https://www.google.com/maps?q="
        f"{latitude},{longitude}"
    )

    map_cell = final_sheet.cell(
        row=row,
        column=MAP_LOCATION_COL
    )

    map_cell.value = (
        "View Map Location"
    )

    map_cell.hyperlink = map_url

    # =====================================================
    # BOLD CUSTOM SCALE
    # =====================================================

    custom_cell = final_sheet.cell(
        row=row,
        column=CUSTOM_SCALE_COL
    )

    custom_font = copy(
        custom_cell.font
    )

    custom_font.bold = True

    custom_cell.font = custom_font

    # =====================================================
    # MAP LOCATION FONT
    # =====================================================

    map_font = copy(
        custom_cell.font
    )

    map_font.bold = True

    map_cell.font = map_font

# =========================================================
# COLUMN WIDTH
# =========================================================

final_sheet.column_dimensions[
    final_sheet.cell(
        row=1,
        column=MAP_LOCATION_COL
    ).column_letter
].width = 22

print(
    "\n✅ TASK 9 COMPLETED\n"
)
# =========================================================
# TASK 10
# SORT FINAL BY CUSTOM SCALE
# =========================================================

print("\n🚀 TASK 10 STARTED\n")

# =========================================================
# FIND CUSTOM SCALE COLUMN
# =========================================================

headers = [
    cell.value
    for cell in final_sheet[1]
]

header_index = {
    header: index + 1
    for index, header in enumerate(headers)
}

CUSTOM_SCALE_COL = header_index[
    CUSTOM_SCALE_COL_NAME
]

# =========================================================
# READ ALL DATA ROWS
# =========================================================

all_rows = []

for row in range(
    2,
    final_sheet.max_row + 1
):

    row_data = []

    for col in range(
        1,
        final_sheet.max_column + 1
    ):

        cell = final_sheet.cell(
            row=row,
            column=col
        )

        row_data.append({

            "value": cell.value,
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "number_format": cell.number_format,
            "hyperlink": (
                cell.hyperlink.target
                if cell.hyperlink
                else None
            )

        })

    all_rows.append(row_data)

# =========================================================
# SORT BY CUSTOM SCALE
# =========================================================

all_rows.sort(
    key=lambda row:
    str(
        row[
            CUSTOM_SCALE_COL - 1
        ]["value"]
    )
)

# =========================================================
# CLEAR OLD DATA
# =========================================================

if final_sheet.max_row > 1:

    final_sheet.delete_rows(
        2,
        final_sheet.max_row
    )

# =========================================================
# WRITE SORTED DATA
# =========================================================

write_row = 2

for row_data in all_rows:

    for col, item in enumerate(
        row_data,
        start=1
    ):

        cell = final_sheet.cell(
            row=write_row,
            column=col
        )

        cell.value = item["value"]
        cell.font = copy(
            item["font"]
        )
        cell.fill = copy(
            item["fill"]
        )
        cell.border = copy(
            item["border"]
        )
        cell.alignment = copy(
            item["alignment"]
        )
        cell.number_format = (
            item["number_format"]
        )

        if item["hyperlink"]:
            cell.hyperlink = (
                item["hyperlink"]
            )

    write_row += 1

print(
    "\n✅ TASK 10 COMPLETED\n"
)
# =========================================================
# SAVE
# =========================================================

workbook.save(OUTPUT_FILE)


print(f"📁 OUTPUT FILE → {OUTPUT_FILE}")
